import pandas as pd
from openpyxl import load_workbook
# CSVファイルを読み込みます
csv_file = "exams.csv"  # ファイル名を適切なものに変更してください
df = pd.read_csv(csv_file)

# エクセルファイルに変換して保存します
excel_file = "output.xlsx"  # 保存するエクセルファイルの名前
df.to_excel(excel_file, index=False)  # index=Falseで行番号を含めないようにします
# エクセルファイルを読み込みます
excel_file = "output.xlsx"  # ファイル名を適切なものに変更してください
wb = load_workbook(excel_file)
ws = wb.active

# 新しいカラムのヘッダーとデータを定義します
new_column_header = "average"
new_column_data = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]  # 適切なデータを指定してください

# 新しいカラムを追加します
ws.insert_cols(ws.max_column + 1)
ws.cell(row=1, column=ws.max_column, value=new_column_header)
for i, data in enumerate(new_column_data, start=2):
    ws.cell(row=i, column=ws.max_column, value=data)

# 変更を保存します
wb.save("output.xlsx")  # 保存するエクセルファイルの名前を適切なものに変更してください